import openpyxl  # Убедитесь, что у вас установлена библиотека openpyxl

class StudentGrades:
    def __init__(self, grades, students):
        self.grades = {}
        for (student_id, student_ref), grade in grades.items():
            student_name, student_st_id = list(students[student_ref].items())[0]  # Извлекаем имя и ID студента
            self.grades[student_name] = (student_st_id, grade)  # Используем student_st_id из students

    def to_excel(self, filename):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Students's Grades"

        ws['A1'] = "St ID"
        ws['B1'] = "Name Surname"
        ws['C1'] = "Grade"

        for idx, (name, (st_id, grade)) in enumerate(self.grades.items(), start=2):
            ws[f'A{idx}'] = st_id  # Записываем ID студента из students
            ws[f'B{idx}'] = name
            ws[f'C{idx}'] = grade

        wb.save(filename)

# Пример использования
# grades = {(12345, 2): 87, (23456, 4): 93}
# students = {2: {'John Doe': 54321}, 4: {'Jane Smith': 34256}}
# report = StudentGrades(grades, students)
# report.to_excel(r'C:\Users\user\Desktop\test\example.xlsx')



# class StudentGrades:
#     def __init__(self, grades, students):
#         self.grades = {students[key]: value for key, value in grades.items()}

#     def to_excel(self, filename):
#         wb = openpyxl.Workbook()
#         ws = wb.active
#         ws.title = "Students's Grades"
        
#         ws['A1'] = "Name Surname"
#         ws['B1'] = "Grade"
        
#         for idx, (name, grade) in enumerate(self.grades.items(), start=2):
#             ws[f'A{idx}'] = name
#             ws[f'B{idx}'] = grade
            
#         wb.save(filename)